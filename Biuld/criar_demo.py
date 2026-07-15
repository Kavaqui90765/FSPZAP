import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Atendimentos"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Cabeçalhos
headers = [
    "ORIGEM", "PROTOCOLO", "STATUS", "NUMERO", "DATA",
    "ATENDENTE", "DEPARTAMENTO", "MOTIVO", "NOME", "DATA_FINALIZACAO",
    "DATA_ULTIMA_MENSAGEM", "POSUI_ANEXO", "AVALIACAO"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Largura das colunas
column_widths = [12, 15, 12, 15, 20, 18, 15, 25, 20, 20, 20, 12, 10]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Dados demo
atendentes = ["Maria Silva", "João Santos", "Ana Oliveira", "Pedro Costa", "Lucia Ferreira"]
departamentos = ["Suporte", "Vendas", "Financeiro", "Técnico", "Comercial"]
motivos = [
    "Dúvida sobre produto", "Problema com pedido", "Solicitação de orçamento",
    "Reclamação", "Informação sobre prazo", "Sugestão de melhoria",
    "Cancelar assinatura", "Alterar dados cadastrais", "Relatar bug",
    "Solicitar demo", "Dúvida sobre faturamento", "Elogio ao atendimento"
]
nomes = [
    "Carlos Eduardo", "Fernanda Lima", "Roberto Almeida", "Mariana Costa",
    "Lucas Ferreira", "Juliana Martins", "Ricardo Souza", "Patricia Araujo",
    "Marcos Ribeiro", "Camila Rodrigues", "Thiago Gomes", "Isabela Dias",
    "Felipe Barbosa", "Amanda Nascimento", "Bruno Carvalho", "Letícia Melo"
]
origens = ["WhatsApp", "Telefone", "E-mail", "Chat", "Instagram"]
statuses = ["Finalizado", "Em andamento", "Pendente", "Transferido"]

# Gerar 50 registros
base_date = datetime(2026, 7, 1)
for i in range(50):
    row = i + 2
    
    # Data aleatória no mês de julho
    days_offset = random.randint(0, 10)
    hours = random.randint(8, 18)
    minutes = random.randint(0, 59)
    data_atendimento = base_date + timedelta(days=days_offset, hours=hours, minutes=minutes)
    data_finalizacao = data_atendimento + timedelta(minutes=random.randint(5, 120))
    data_ultima_msg = data_finalizacao - timedelta(minutes=random.randint(0, 30))
    
    status = random.choice(statuses)
    protocolo = f"PROTO-{2026}{i+1:04d}"
    numero = f"+5511{random.randint(900000000, 999999999)}"
    
    ws.cell(row=row, column=1, value=random.choice(origens))
    ws.cell(row=row, column=2, value=protocolo)
    ws.cell(row=row, column=3, value=status)
    ws.cell(row=row, column=4, value=numero)
    ws.cell(row=row, column=5, value=data_atendimento.strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row, column=6, value=random.choice(atendentes))
    ws.cell(row=row, column=7, value=random.choice(departamentos))
    ws.cell(row=row, column=8, value=random.choice(motivos))
    ws.cell(row=row, column=9, value=random.choice(nomes))
    ws.cell(row=row, column=10, value=data_finalizacao.strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row, column=11, value=data_ultima_msg.strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row, column=12, value=random.choice([0, 1]))
    ws.cell(row=row, column=13, value=random.choice([None, 1, 2, 3, 4, 5]))
    
    # Bordas e alinhamento
    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if col in [1, 3, 12, 13]:  # Centralizar colunas específicas
            cell.alignment = Alignment(horizontal="center")

# Congelar cabeçalho
ws.freeze_panes = "A2"

# Auto-filtro
ws.auto_filter.ref = f"A1:M{51}"

# Salvar
output_path = r"D:\Projetos\MEGAZAP\demo_fspzap.xlsx"
wb.save(output_path)
print(f"Planilha demo criada: {output_path}")
print(f"Total de registros: 50")
